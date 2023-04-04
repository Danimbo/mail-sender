import smtplib

from openpyxl import load_workbook
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import os
import mimetypes

mails = []

wb = load_workbook(filename = 'base.xlsx')
ws = wb.active

for row in ws.iter_rows(min_row=5, min_col=8, max_col=8, max_row=8, values_only=True):
    for iter_row in row:
        if iter_row != None:
            str(iter_row)
            mails.append(iter_row)
            # print(iter_row)
    

def send_email(message, receiver):
    sender = "lomex2014@mail.ru"
    # your password = "your password"
    password = "x3cy6Hw2kbwC7qSp4d7y"
    # brugrmgulsghtwok
    server = smtplib.SMTP_SSL("smtp.mail.ru", 465)
    print(receiver)
 
    try:
        server.login(sender, password)
        # msg = MIMEText(message)
        msg = MIMEMultipart()
        msg["Subject"] = "Офсетная печать по низким ценам"
        


        msg.attach(MIMEText(message))

        for file in os.listdir("attachments"):
            filename = os.path.basename(file)
            ftype, encoding = mimetypes.guess_type(file)
            file_type, subtype = ftype.split("/")
            print(file_type, subtype)

            if file_type == "application":
                with open(f"attachments/{file}", "rb") as f:
                    file = MIMEApplication(f.read(), subtype)

        file.add_header('content-disposition', 'attachment', filename=filename)
        msg.attach(file)

        server.sendmail(sender, receiver, msg.as_string())

        return "The message was sent successfully!"
    except Exception as _ex:
        return f"{_ex}\nCheck your login or password please!"

print(len(mails))



for mail in mails:
        message = """Рекламное Агентство ХХI век полного цикла предлагает услуги офсетной печати визиток, листовок 
и буклетов по низким ценам. 

Мы предоставляем лучшие решения для печати, которые осуществляем по современным 
технологиям и стандартам. 
                     
Разрабатываем креативный дизайн рекламной продукции.

Быстрый и удобный формат заказа онлайн.

Нам доверяют сотни довольных клиентов: Администрация города Мурманска,  Ростелеком, 
КОЛАТОМЭНЕРГОСБЫТ, МТС, НМКК ФОРМАП, ГОБУ МРИБИ, Роснефть, Почта РФ, ГАЗПРОМ, ГОБУЗ 
МОКМЦ, Хлебопек, МГТУ, ТЦ Мегастрой, ТЦ Форум, мебельный салон Лазурит, салон красоты Тай-
Рай, ПУЛЬС и многие другие.

Во вложении коммерческое предложение с полным пакетом наших услуг, а также прайсом.
Надеемся на долгосрочное сотрудничество.

С уважением, Сергеева Светлана Викторовна
директор ООО "Рекламное Агентство ХХI век"
183036, г. Мурманск, ТЦ М-СИТИ, ул. Старостина, д.55,  2 этаж
моб.:+7 911 315 27 42
ra-21vek.com
vk.com/rk21vek"""
        print(send_email(message=message, receiver=str(mail)))
        
        


